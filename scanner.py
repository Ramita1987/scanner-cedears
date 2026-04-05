"""
╔══════════════════════════════════════════════════════════════╗
║         SCANNER PROFESIONAL DE CEDEARS - v1.0               ║
║         Intraday + Swing | Daily + 1H | Telegram Alerts     ║
╚══════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import pandas as pd
import numpy as np
import time
import logging
from datetime import datetime
from typing import Optional
import requests
import sys
import os
from urllib.parse import quote

# ── Importar configuración ──────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import TELEGRAM_TOKEN, TELEGRAM_CHAT_ID, TOP_N, MIN_PROBABILITY, CEDEARS

FMP_KEY = os.environ.get("FMP_KEY", "")
GOOGLE_SHEETS_WEBHOOK_URL = os.environ.get("GOOGLE_SHEETS_WEBHOOK_URL", "").strip()

# ── Logger ──────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(os.path.join(os.path.dirname(__file__), "scanner.log")),
    ],
)
log = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
#  DESCARGA DE DATOS
# ═══════════════════════════════════════════════════════════════

def _fmp_daily_ohlc(ticker: str) -> Optional[pd.DataFrame]:
    """Fallback diario via FMP para cuando Yahoo falla."""
    if not FMP_KEY:
        return None
    try:
        enc = quote(ticker, safe="")
        url = (
            f"https://financialmodelingprep.com/stable/historical-price-eod/full"
            f"?symbol={enc}&apikey={FMP_KEY}"
        )
        r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code != 200:
            return None
        payload = r.json()
        if isinstance(payload, list):
            rows = payload
        elif isinstance(payload, dict):
            rows = payload.get("historical", [])
        else:
            rows = []
        if not rows:
            return None
        df = pd.DataFrame(rows)
        req = {"date", "open", "high", "low", "close", "volume"}
        if not req.issubset(df.columns):
            return None
        df = df[["date", "open", "high", "low", "close", "volume"]].copy()
        df["date"] = pd.to_datetime(df["date"])
        df.sort_values("date", inplace=True)
        df.set_index("date", inplace=True)
        for col in ["open", "high", "low", "close", "volume"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        df.dropna(subset=["close", "volume"], inplace=True)
        return df
    except Exception:
        return None


def _normalize_ohlc(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [c.lower() for c in df.columns]
    df.index = pd.to_datetime(df.index)
    df.dropna(subset=["close", "volume"], inplace=True)
    return df


def fetch_data_resilient(ticker: str) -> tuple:
    """VersiÃ³n robusta: Yahoo + fallback FMP, y 1H opcional."""
    try:
        t = yf.Ticker(ticker)
        df_d = t.history(period="6mo", interval="1d", auto_adjust=True)
        df_h = t.history(period="60d", interval="1h", auto_adjust=True)
    except Exception as e:
        log.debug(f"{ticker} - error descarga yahoo: {e}")
        df_d, df_h = pd.DataFrame(), pd.DataFrame()

    if df_d is not None and not df_d.empty:
        df_d = _normalize_ohlc(df_d)
    else:
        df_d = pd.DataFrame()

    if df_h is not None and not df_h.empty:
        df_h = _normalize_ohlc(df_h)
    else:
        df_h = pd.DataFrame()

    if df_d.empty or len(df_d) < 30:
        df_fmp = _fmp_daily_ohlc(ticker)
        if df_fmp is not None and len(df_fmp) >= 30:
            df_d = _normalize_ohlc(df_fmp)

    if df_d.empty or len(df_d) < 30:
        return None, None

    if df_h.empty or len(df_h) < 20:
        df_h = df_d.copy()

    return df_d, df_h


def fetch_data(ticker: str) -> tuple:
    """
    Descarga datos Daily (6 meses) y 1H (60 días) para un ticker.
    Retorna (df_daily, df_1h) o (None, None) si falla.
    """
    try:
        t = yf.Ticker(ticker)

        df_d = t.history(period="6mo", interval="1d", auto_adjust=True)
        df_h = t.history(period="60d",  interval="1h", auto_adjust=True)

        if df_d.empty or len(df_d) < 30:
            return None, None
        if df_h.empty or len(df_h) < 20:
            return None, None

        # Normalizar columnas
        for df in [df_d, df_h]:
            df.columns = [c.lower() for c in df.columns]
            df.index = pd.to_datetime(df.index)
            df.dropna(subset=["close", "volume"], inplace=True)

        return df_d, df_h

    except Exception as e:
        log.debug(f"{ticker} — error descarga: {e}")
        return None, None


# ═══════════════════════════════════════════════════════════════
#  INDICADORES TÉCNICOS
# ═══════════════════════════════════════════════════════════════

def add_indicators(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega EMA20, SMA50, SMA200, RSI14, ATR14, VolAvg20 al DataFrame."""
    df = df.copy()
    c = df["close"]
    h = df["high"]
    l = df["low"]

    # Medias móviles
    df["ema20"]  = c.ewm(span=20, adjust=False).mean()
    df["sma50"]  = c.rolling(50).mean()
    df["sma200"] = c.rolling(200).mean()

    # RSI 14
    delta = c.diff()
    gain  = delta.clip(lower=0).rolling(14).mean()
    loss  = (-delta.clip(upper=0)).rolling(14).mean()
    rs    = gain / loss.replace(0, np.nan)
    df["rsi"] = 100 - (100 / (1 + rs))

    # ATR 14
    tr = pd.concat([
        h - l,
        (h - c.shift()).abs(),
        (l - c.shift()).abs()
    ], axis=1).max(axis=1)
    df["atr"] = tr.rolling(14).mean()

    # Volumen promedio 20 días
    df["vol_avg20"] = df["volume"].rolling(20).mean()

    return df


# ═══════════════════════════════════════════════════════════════
#  DETECCIÓN DE SETUPS
# ═══════════════════════════════════════════════════════════════

def detect_setups(df_d: pd.DataFrame, df_h: pd.DataFrame) -> list:
    """
    Evalúa los 7 setups sobre los datos.
    Retorna lista de setups detectados con metadata.
    """
    setups = []
    d  = df_d.iloc[-1]
    d1 = df_d.iloc[-2]

    close   = d["close"]
    rsi_d   = d["rsi"]
    ema20_d = d["ema20"]
    sma50_d = d["sma50"]
    vol_rel = d["volume"] / d["vol_avg20"] if d["vol_avg20"] > 0 else 1.0

    # ── SETUP 1 — Pullback a EMA20 ─────────────────────────────
    try:
        trend_up  = close > sma50_d and d1["close"] > d1["sma50"]
        sobre_ema = close > ema20_d
        pullback  = abs(close - ema20_d) / ema20_d < 0.015
        rsi_ok    = 40 <= rsi_d <= 60
        vela_alc  = d["close"] > d["open"]

        if trend_up and sobre_ema and pullback and rsi_ok and vela_alc:
            setups.append({
                "setup":       "Pullback EMA20",
                "calidad":     _calidad_pullback(df_d),
                "descripcion": "Tendencia alcista, precio toca EMA20 con RSI neutro"
            })
    except Exception:
        pass

    # ── SETUP 2 — RSI Mean Reversion ───────────────────────────
    try:
        rsi_ob = rsi_d < 32
        if rsi_ob:
            setups.append({
                "setup":       "RSI Mean Reversion",
                "calidad":     5 if rsi_d < 25 else 4 if rsi_d < 30 else 3,
                "descripcion": f"RSI en sobreventa ({rsi_d:.1f}), rebote técnico esperado"
            })
    except Exception:
        pass

    # ── SETUP 3 — Gap Fill ─────────────────────────────────────
    try:
        gap_up   = d["open"] > d1["high"] * 1.005
        gap_down = d["open"] < d1["low"]  * 0.995

        if gap_up or gap_down:
            ref     = d1["high"] if gap_up else d1["low"]
            gap_pct = abs(d["open"] - ref) / d1["close"] * 100
            if gap_pct > 0.5:
                setups.append({
                    "setup":       "Gap Fill",
                    "calidad":     min(5, max(2, int(gap_pct))),
                    "descripcion": f"Gap {'alcista' if gap_up else 'bajista'} de {gap_pct:.1f}% detectado"
                })
    except Exception:
        pass

    # ── SETUP 4 — Breakout de Resistencia ──────────────────────
    try:
        resistencia = df_d["high"].iloc[-20:-1].max()
        breakout    = close > resistencia * 1.005
        vol_alto    = vol_rel > 1.5

        if breakout and vol_alto:
            setups.append({
                "setup":       "Breakout Resistencia",
                "calidad":     5 if vol_rel > 2.5 else 4 if vol_rel > 2.0 else 3,
                "descripcion": f"Ruptura de resistencia {resistencia:.2f} con volumen {vol_rel:.1f}x"
            })
    except Exception:
        pass

    # ── SETUP 5 — Rebote en Soporte ────────────────────────────
    try:
        soporte      = df_d["low"].iloc[-20:-1].min()
        cerca_sop    = abs(close - soporte) / soporte < 0.02
        vela_rechazo = d["close"] > d["open"] and d["low"] < d1["low"]

        if cerca_sop and vela_rechazo:
            setups.append({
                "setup":       "Rebote en Soporte",
                "calidad":     _calidad_soporte(df_d),
                "descripcion": f"Precio rebota en soporte {soporte:.2f} con vela de rechazo"
            })
    except Exception:
        pass

    # ── SETUP 6 — Compresión de Volatilidad ────────────────────
    try:
        atrs = df_d["atr"].iloc[-10:]
        if len(atrs) >= 10:
            atr_rec = atrs.iloc[-3:].mean()
            atr_ant = atrs.iloc[:7].mean()
            if atr_ant > 0 and atr_rec < atr_ant * 0.65:
                setups.append({
                    "setup":       "Compresión Volatilidad",
                    "calidad":     4 if atr_rec < atr_ant * 0.5 else 3,
                    "descripcion": f"ATR comprimido al {atr_rec/atr_ant*100:.0f}% — posible expansión próxima"
                })
    except Exception:
        pass

    # ── SETUP 7 — Reversión Caída Vertical ─────────────────────
    try:
        if len(df_d) >= 5:
            caida_3d   = (df_d["close"].iloc[-4] - close) / df_d["close"].iloc[-4]
            rsi_ob     = rsi_d < 35
            primer_reb = d["close"] > d1["close"]

            if caida_3d > 0.06 and rsi_ob and primer_reb:
                setups.append({
                    "setup":       "Reversión Caída Vertical",
                    "calidad":     4 if caida_3d > 0.10 else 3,
                    "descripcion": f"Caída de {caida_3d*100:.1f}% en 3 días, primer rebote RSI {rsi_d:.1f}"
                })
    except Exception:
        pass

    return setups


def _calidad_pullback(df: pd.DataFrame) -> int:
    highs = df["high"].iloc[-5:]
    lows  = df["low"].iloc[-5:]
    if highs.is_monotonic_increasing and lows.is_monotonic_increasing:
        return 5
    if highs.iloc[-1] > highs.iloc[0]:
        return 4
    return 3


def _calidad_soporte(df: pd.DataFrame) -> int:
    soporte = df["low"].iloc[-20:-1].min()
    toques  = ((df["low"].iloc[-20:] - soporte).abs() / soporte < 0.02).sum()
    return min(5, max(2, int(toques)))


# ═══════════════════════════════════════════════════════════════
#  SCORE DE CONFLUENCIAS
# ═══════════════════════════════════════════════════════════════

def calc_confluencias(df_d: pd.DataFrame) -> tuple:
    """Evalúa 5 factores (1 punto c/u). Retorna (score, detalle)."""
    score   = 0
    detalle = {}
    d = df_d.iloc[-1]
    c = d["close"]

    # Factor 1 — Tendencia
    max_crec  = df_d["high"].iloc[-5:].is_monotonic_increasing
    min_crec  = df_d["low"].iloc[-5:].is_monotonic_increasing
    sobre_sma = c > d["sma50"] if not np.isnan(d["sma50"]) else False
    f1 = 1 if (sobre_sma and (max_crec or min_crec)) else 0
    score += f1; detalle["tendencia"] = f1

    # Factor 2 — Soporte/Resistencia
    soporte     = df_d["low"].iloc[-20:-1].min()
    resistencia = df_d["high"].iloc[-20:-1].max()
    f2 = 1 if (abs(c - soporte) / soporte < 0.03 or abs(c - resistencia) / resistencia < 0.03) else 0
    score += f2; detalle["soporte_resistencia"] = f2

    # Factor 3 — Media Móvil
    ema20  = d["ema20"]
    sma50  = d["sma50"]  if not np.isnan(d["sma50"])  else None
    sma200 = d["sma200"] if not np.isnan(d["sma200"]) else None
    t_ema  = abs(c - ema20)  / ema20  < 0.02
    t_50   = abs(c - sma50)  / sma50  < 0.02 if sma50  else False
    t_200  = abs(c - sma200) / sma200 < 0.02 if sma200 else False
    f3 = 1 if (t_ema or t_50 or t_200) else 0
    score += f3; detalle["media_movil"] = f3

    # Factor 4 — RSI
    rsi = d["rsi"]
    f4  = 1 if (rsi < 35 or (40 <= rsi <= 60)) else 0
    score += f4; detalle["rsi"] = f4

    # Factor 5 — Volumen
    vol_rel = d["volume"] / d["vol_avg20"] if d["vol_avg20"] > 0 else 1.0
    f5 = 1 if vol_rel > 1.3 else 0
    score += f5; detalle["volumen"] = f5

    return score, detalle


# ═══════════════════════════════════════════════════════════════
#  FÓRMULA DE PROBABILIDAD
# ═══════════════════════════════════════════════════════════════

def calc_probabilidad(score_conf: int, calidad_setup: int, df_d: pd.DataFrame) -> float:
    d = df_d.iloc[-1]

    vol_rel_raw = d["volume"] / d["vol_avg20"] if d["vol_avg20"] > 0 else 1.0
    vol_score   = float(np.clip(vol_rel_raw * 1.5, 1, 5))

    rsi = d["rsi"]
    if   rsi < 25:            rsi_score = 5
    elif rsi < 30:            rsi_score = 4
    elif rsi < 35:            rsi_score = 3
    elif 40 <= rsi <= 60:     rsi_score = 3
    elif rsi > 70:            rsi_score = 1
    else:                     rsi_score = 2

    conf_norm = (score_conf / 5) * 5

    raw = (
        0.35 * conf_norm    +
        0.30 * calidad_setup +
        0.20 * vol_score    +
        0.15 * rsi_score
    )

    prob = ((raw - 1) / 4) * 100
    return round(float(np.clip(prob, 0, 100)), 1)


# ═══════════════════════════════════════════════════════════════
#  ANÁLISIS POR TICKER
# ═══════════════════════════════════════════════════════════════

def analizar_ticker(ticker: str, active_setups: list = None) -> Optional[dict]:
    df_d, df_h = fetch_data_resilient(ticker)
    if df_d is None:
        return None

    df_d = add_indicators(df_d)
    df_h = add_indicators(df_h)

    setups = detect_setups(df_d, df_h)

    # Filtrar por setups activos si se especificaron
    if active_setups:
        setups = [s for s in setups if s["setup"] in active_setups]

    if not setups:
        return None

    score_conf, detalle_conf = calc_confluencias(df_d)
    if score_conf < 2:
        return None

    mejor_setup = max(setups, key=lambda s: s["calidad"])
    prob        = calc_probabilidad(score_conf, mejor_setup["calidad"], df_d)

    d       = df_d.iloc[-1]
    vol_rel = d["volume"] / d["vol_avg20"] if d["vol_avg20"] > 0 else 1.0

    return {
        "ticker":       ticker.replace(".BA", ""),
        "ticker_full":  ticker,
        "setup":        mejor_setup["setup"],
        "confluencias": score_conf,
        "probabilidad": prob,
        "calidad":      mejor_setup["calidad"],
        "descripcion":  mejor_setup["descripcion"],
        "precio":       round(float(d["close"]), 2),
        "rsi":          round(float(d["rsi"]), 1),
        "vol_rel":      round(float(vol_rel), 2),
        "atr":          round(float(d["atr"]), 2),
        "detalle_conf": detalle_conf,
        "setups_extra": [s["setup"] for s in setups if s["setup"] != mejor_setup["setup"]],
    }


# ═══════════════════════════════════════════════════════════════
#  RUNNER PRINCIPAL
# ═══════════════════════════════════════════════════════════════

def run_scanner(session_name: str = "MANUAL") -> list:
    log.info("=" * 60)
    log.info(f"  SCANNER INICIADO — Sesión: {session_name}")
    log.info(f"  Hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info(f"  Tickers a analizar: {len(CEDEARS)}")
    log.info("=" * 60)

    resultados = []
    errores    = 0

    for i, ticker in enumerate(CEDEARS, 1):
        log.info(f"[{i:>3}/{len(CEDEARS)}] {ticker}...")
        try:
            res = analizar_ticker(ticker)
            if res and res["probabilidad"] >= MIN_PROBABILITY:
                resultados.append(res)
                log.info(f"         ✅ {res['setup']} | {res['probabilidad']}% | {res['confluencias']}/5")
        except Exception as e:
            log.warning(f"         ⚠️  {e}")
            errores += 1

        # Rate limiting
        if i % 10 == 0:
            time.sleep(2)
        else:
            time.sleep(0.3)

    # Ranking
    resultados.sort(
        key=lambda x: (x["probabilidad"], x["confluencias"], x["vol_rel"]),
        reverse=True
    )
    top = resultados[:TOP_N]

    log.info(f"\n  Completado. Oportunidades: {len(resultados)} | Errores: {errores}")
    return top


# ═══════════════════════════════════════════════════════════════
#  TELEGRAM
# ═══════════════════════════════════════════════════════════════

CONF_STARS = {5: "⭐⭐⭐⭐⭐", 4: "⭐⭐⭐⭐", 3: "⭐⭐⭐", 2: "⭐⭐", 1: "⭐"}


def prob_emoji(p: float) -> str:
    return "🔥" if p >= 80 else "✅" if p >= 70 else "🟡"


def build_telegram_message(oportunidades: list, session_name: str) -> str:
    hora  = datetime.now().strftime("%H:%M")
    fecha = datetime.now().strftime("%d/%m/%Y")
    emojis = {"PRE-MARKET": "🌅", "APERTURA": "🔔", "CIERRE": "🌙", "MANUAL": "🔍"}
    emoji  = emojis.get(session_name, "🔍")

    header = (
        f"{emoji} *SCANNER CEDEARS — {session_name}*\n"
        f"📅 {fecha}  🕐 {hora} (ARG)\n"
        f"{'─' * 32}\n\n"
    )

    if not oportunidades:
        return header + "😴 *Sin oportunidades claras en este momento.*\nNo hay setups con probabilidad > 60%."

    body = ""
    for i, op in enumerate(oportunidades, 1):
        extras = f" _(+{', '.join(op['setups_extra'])})_" if op["setups_extra"] else ""

        # Armar detalle de confluencias
        conf_detalle = []
        dc = op["detalle_conf"]
        if dc.get("tendencia"):           conf_detalle.append("📈 Tendencia")
        if dc.get("soporte_resistencia"): conf_detalle.append("🧱 Soporte/Res")
        if dc.get("media_movil"):         conf_detalle.append("〰️ Media Móvil")
        if dc.get("rsi"):                 conf_detalle.append("📉 RSI")
        if dc.get("volumen"):             conf_detalle.append("📊 Volumen")
        conf_texto = " | ".join(conf_detalle) if conf_detalle else "—"

        body += (
            f"*{i}. {op['ticker']}* — {prob_emoji(op['probabilidad'])} `{op['probabilidad']}%`\n"
            f"   📐 Setup: *{op['setup']}*{extras}\n"
            f"   {CONF_STARS.get(op['confluencias'], '⭐')} Confluencias: {op['confluencias']}/5\n"
            f"   ✅ _{conf_texto}_\n"
            f"   💲 Precio: `${op['precio']}`  |  RSI: `{op['rsi']}`\n"
            f"   📊 Vol: `{op['vol_rel']}x`  |  ATR: `{op['atr']}`\n"
            f"   💬 _{op['descripcion']}_\n\n"
        )

    footer = (
        f"{'─' * 32}\n"
        f"📚 *Guía de Setups:*\n"
        f"1️⃣ *Pullback EMA20* — Tendencia alcista, precio retrocede a EMA20 con RSI neutro. Buscá vela verde de rebote.\n"
        f"2️⃣ *RSI Mean Reversion* — RSI bajo 32, sobreventa extrema. Esperá primera vela verde como confirmación.\n"
        f"3️⃣ *Gap Fill* — Hueco entre velas. El precio tiende a volver a taparlo. Confirmá dirección del gap.\n"
        f"4️⃣ *Breakout Resistencia* — Precio rompe máximos con volumen alto. Entrada en pullback post-ruptura.\n"
        f"5️⃣ *Rebote en Soporte* — Precio llega a zona de soporte probada. Buscá mecha larga abajo y cierre arriba.\n"
        f"6️⃣ *Compresión Volatilidad* — Velas cada vez más chicas. Preparate para movimiento fuerte próximo.\n"
        f"7️⃣ *Reversión Caída Vertical* — Caída brusca seguida de primer rebote. RSI en sobreventa confirma.\n"
        f"{'─' * 32}\n"
        f"⚠️ _Confirmá siempre en el gráfico antes de entrar._\n"
        f"🤖 Scanner automático — no es recomendación financiera."
    )

    return header + body + footer


def send_telegram(message: str) -> bool:
    if not TELEGRAM_TOKEN or TELEGRAM_TOKEN == "TU_TOKEN_AQUI":
        log.warning("Telegram no configurado — omitiendo envío.")
        print("\n" + "─" * 50)
        print("MENSAJE QUE SE ENVIARÍA A TELEGRAM:")
        print("─" * 50)
        print(message)
        return False

    url     = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": message, "parse_mode": "Markdown"}

    try:
        r = requests.post(url, json=payload, timeout=15)
        if r.status_code == 200:
            log.info("✅ Mensaje enviado a Telegram.")
            return True
        else:
            log.error(f"❌ Error Telegram {r.status_code}: {r.text[:200]}")
            return False
    except Exception as e:
        log.error(f"❌ Excepción Telegram: {e}")
        return False


# ═══════════════════════════════════════════════════════════════
#  REGISTRO EXCEL
# ═══════════════════════════════════════════════════════════════

def _send_google_sheets_rows(rows: list) -> None:
    """
    Envia filas a un webhook (Apps Script) para guardar en Google Sheets.
    Espera la variable de entorno GOOGLE_SHEETS_WEBHOOK_URL.
    """
    if not GOOGLE_SHEETS_WEBHOOK_URL or not rows:
        return
    try:
        r = requests.post(GOOGLE_SHEETS_WEBHOOK_URL, json={"rows": rows}, timeout=15)
        if 200 <= r.status_code < 300:
            log.info(f"Google Sheets actualizado ({len(rows)} filas).")
        else:
            log.warning(f"Google Sheets webhook {r.status_code}: {r.text[:180]}")
    except Exception as e:
        log.warning(f"No se pudo enviar a Google Sheets: {e}")


def guardar_excel(oportunidades: list, session_name: str):
    """Guarda las oportunidades en un Excel acumulativo."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill

        archivo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "registros_scanner.xlsx")
        fecha   = datetime.now().strftime("%d/%m/%Y")
        hora    = datetime.now().strftime("%H:%M")

        if os.path.exists(archivo):
            wb = load_workbook(archivo)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Registros"
            ws.append([
                "Fecha", "Hora", "Sesión", "Ticker", "Setup",
                "Confluencias", "Probabilidad %", "Precio Entrada",
                "Target +10%", "Stop -5%", "RSI",
                "Volumen Rel", "ATR", "Descripción", "Resultado"
            ])
            from openpyxl.styles import Font, PatternFill
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")

        gs_rows = []
        for op in oportunidades:
            precio  = op["precio"]
            target  = round(precio * 1.10, 2)
            stop    = round(precio * 0.95, 2)
            ws.append([
                fecha, hora, session_name,
                op["ticker"], op["setup"],
                op["confluencias"], op["probabilidad"],
                precio, target, stop,
                op["rsi"], op["vol_rel"], op["atr"],
                op["descripcion"], "Pendiente"
            ])
            gs_rows.append({
                "fecha": fecha,
                "hora": hora,
                "sesion": session_name,
                "ticker": op["ticker"],
                "setup": op["setup"],
                "confluencias": op["confluencias"],
                "probabilidad": op["probabilidad"],
                "precio": precio,
                "target": target,
                "stop": stop,
                "rsi": op["rsi"],
                "vol_rel": op["vol_rel"],
                "atr": op["atr"],
                "descripcion": op["descripcion"],
                "resultado": "Pendiente",
            })

        wb.save(archivo)
        _send_google_sheets_rows(gs_rows)
        log.info(f"✅ Excel actualizado: {archivo}")

    except Exception as e:
        log.error(f"❌ Error guardando Excel: {e}")


# ═══════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def main(session: str = "MANUAL"):
    session = session.upper()

    oportunidades = run_scanner(session)

    # Output en consola
    print("\n" + "=" * 62)
    print(f"  TOP {TOP_N} OPORTUNIDADES — Sesión: {session}")
    print("=" * 62)
    if oportunidades:
        print(f"{'#':<3} {'Ticker':<8} {'Setup':<26} {'Conf':>5} {'Prob%':>6} {'Vol':>6} {'RSI':>6}")
        print("─" * 62)
        for i, op in enumerate(oportunidades, 1):
            print(
                f"{i:<3} {op['ticker']:<8} {op['setup']:<26} "
                f"{op['confluencias']:>5} {op['probabilidad']:>5.1f}% "
                f"{op['vol_rel']:>5.1f}x {op['rsi']:>5.1f}"
            )
    else:
        print("  Sin oportunidades con probabilidad > 60%")
    print("=" * 62 + "\n")

    # Enviar Telegram
    msg = build_telegram_message(oportunidades, session)
    send_telegram(msg)

    # Guardar en Excel
    if oportunidades:
        guardar_excel(oportunidades, session)

    return oportunidades


if __name__ == "__main__":
    session_arg = sys.argv[1] if len(sys.argv) > 1 else "MANUAL"
    main(session_arg)
