"""
╔══════════════════════════════════════════════════════════════╗
║         SCANNER CEDEARS — INTERFAZ WEB v2                   ║
║         Flask App para Render.com                           ║
╚══════════════════════════════════════════════════════════════╝
"""

from flask import Flask, render_template, jsonify, request
import threading
import os, sys, time
import json
import requests
import feedparser
from datetime import datetime
from urllib.parse import quote
import re
import yfinance as yf

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)

# ── API Keys ─────────────────────────────────────────────────────
AV_KEY  = os.environ.get("AV_KEY",  "6IFZV2E8RQ6BMJ0L")   # Alpha Vantage
FMP_KEY = os.environ.get("FMP_KEY", "aiQvIiYs0bc5eOheSFHH2c4kmi4lRVhr")                 # Financial Modeling Prep (free)

# ── Cache global ─────────────────────────────────────────────────
_cache = {}
MARKET_SNAPSHOT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "market_last_ok.json")

def get_cache(key, ttl=600):
    """Retorna valor cacheado si no expiró (ttl en segundos)."""
    if key in _cache:
        val, ts = _cache[key]
        if (datetime.now() - ts).seconds < ttl:
            return val
    return None

def set_cache(key, val):
    """Solo cachea si hay datos reales."""
    if val is None:
        return val
    # No cachear listas/dicts vacíos
    if isinstance(val, (list, dict)) and len(val) == 0:
        return val
    _cache[key] = (val, datetime.now())
    return val


def load_market_snapshot() -> dict:
    try:
        if not os.path.exists(MARKET_SNAPSHOT_PATH):
            return {}
        with open(MARKET_SNAPSHOT_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_market_snapshot(payload: dict):
    try:
        if not isinstance(payload, dict) or not payload:
            return
        with open(MARKET_SNAPSHOT_PATH, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
    except Exception:
        pass


def data912_live_usa() -> list:
    """Panel USA live de Data912."""
    cached = get_cache("d912_live_usa", ttl=180)
    if cached:
        return cached
    try:
        r = requests.get(
            "https://data912.com/live/usa_stocks",
            timeout=12,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        if r.status_code != 200:
            return []
        payload = r.json()
        if not isinstance(payload, list):
            return []
        set_cache("d912_live_usa", payload)
        return payload
    except Exception:
        return []


def data912_quote_map(symbols: list) -> dict:
    """Mapa symbol -> {price, chg_pct} usando /live/usa_stocks."""
    out = {}
    if not symbols:
        return out
    wanted = {str(s).strip().upper() for s in symbols if str(s).strip()}
    if not wanted:
        return out
    rows = data912_live_usa()
    for row in rows:
        try:
            sym = str(row.get("symbol", "")).strip().upper()
            if sym not in wanted:
                continue
            price = _to_float(row.get("c"), default=None)
            if price is None:
                continue
            chg = _to_float(row.get("pct_change"), default=0.0)
            out[sym] = {"price": round(price, 2), "chg_pct": round(chg, 2)}
        except Exception:
            continue
    return out


def data912_daily(symbol: str, timeseries: int = 130) -> list:
    """Histórico diario Data912 (cedears -> stocks)."""
    key = f"d912d_{symbol}_{timeseries}"
    cached = get_cache(key, ttl=3600)
    if cached:
        return cached
    sym = (symbol or "").strip().upper()
    if not sym:
        return []
    for endpoint in (f"cedears/{sym}", f"stocks/{sym}"):
        try:
            url = f"https://data912.com/historical/{endpoint}"
            r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
            if r.status_code != 200:
                continue
            payload = r.json()
            if not isinstance(payload, list) or not payload:
                continue
            hist = []
            for row in payload:
                d = row.get("date")
                c = _to_float(row.get("c"), default=None)
                if d and c is not None:
                    hist.append({"date": d, "close": round(c, 2)})
            if len(hist) >= 5:
                set_cache(key, hist[-timeseries:])
                return hist[-timeseries:]
        except Exception:
            continue
    return []

# ── Estado scanner ────────────────────────────────────────────────
scanner_state = {
    "running": False, "session": "", "progress": 0, "total": 0,
    "current_ticker": "", "results": [], "log": [], "last_run": None,
}


# ═══════════════════════════════════════════════════════════════
#  FUENTES DE DATOS — Sin Yahoo Finance
# ═══════════════════════════════════════════════════════════════

def fmp_quote(symbols: list) -> dict:
    """
    Financial Modeling Prep — cotizaciones en tiempo real.
    Plan free: funciona con 'demo' para tickers populares.
    Endpoint: https://financialmodelingprep.com/api/v3/quote/AAPL,MSFT
    """
    result = {}
    if not symbols:
        return result
    try:
        syms_str = ",".join(symbols)
        url = f"https://financialmodelingprep.com/api/v3/quote/{syms_str}?apikey={FMP_KEY}"
        r   = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code != 200:
            return result
        data = r.json()
        if not isinstance(data, list):
            return result
        for item in data:
            sym = item.get("symbol","")
            if sym:
                result[sym] = {
                    "price":   round(float(item.get("price", 0)), 2),
                    "chg_pct": round(float(item.get("changesPercentage", 0)), 2),
                }
    except Exception:
        pass
    return result


def fmp_index_quote(symbol_map: dict) -> dict:
    """
    Cotizaciones de índices via FMP.
    symbol_map: {display_name: fmp_symbol}
    e.g. {"S&P 500": "^GSPC", "NASDAQ": "^IXIC"}
    """
    result = {}
    try:
        syms = list(symbol_map.values())
        data = fmp_quote(syms)
        for name, sym in symbol_map.items():
            if sym in data:
                result[name] = data[sym]
    except Exception:
        pass
    return result


def av_quote_single(symbol: str) -> dict:
    """Alpha Vantage — cotización individual."""
    cached = get_cache(f"av_{symbol}", ttl=600)
    if cached:
        return cached
    try:
        url = (f"https://www.alphavantage.co/query"
               f"?function=GLOBAL_QUOTE&symbol={symbol}&apikey={AV_KEY}")
        r = requests.get(url, timeout=10)
        d = r.json().get("Global Quote", {})
        if not d or "05. price" not in d:
            return {}
        val = {
            "price":   round(float(d["05. price"]), 2),
            "chg_pct": round(float(d["10. change percent"].replace("%","")), 2),
        }
        return set_cache(f"av_{symbol}", val)
    except Exception:
        return {}


def av_daily(symbol: str) -> list:
    """Alpha Vantage — historial diario (~5 meses compact)."""
    cached = get_cache(f"avd_{symbol}", ttl=3600)
    if cached:
        return cached
    try:
        url = (f"https://www.alphavantage.co/query"
               f"?function=TIME_SERIES_DAILY&symbol={symbol}"
               f"&outputsize=compact&apikey={AV_KEY}")
        r    = requests.get(url, timeout=15)
        data = r.json().get("Time Series (Daily)", {})
        if not data:
            return []
        hist = [{"date": d, "close": round(float(v["4. close"]), 2)}
                for d, v in sorted(data.items())]
        return set_cache(f"avd_{symbol}", hist[-130:])
    except Exception:
        return []


def _to_float(value, default=None):
    try:
        if value is None:
            return default
        if isinstance(value, str):
            value = value.replace("%", "").strip()
            if not value:
                return default
        return float(value)
    except Exception:
        return default


def fmp_quote_safe(symbols: list) -> dict:
    """Cotizaciones FMP usando /stable con batch cuando es posible."""
    result = {}
    if not symbols:
        return result
    unique = []
    for sym in symbols:
        s = (sym or "").strip()
        if s and s not in unique:
            unique.append(s)
    headers = {"User-Agent": "Mozilla/5.0"}

    # 1) Batch para acciones/ETFs "simples" (menos requests -> menos 429)
    simple = [s for s in unique if re.match(r"^[A-Z0-9.\-]+$", s)]
    if simple:
        try:
            sym_csv = ",".join(simple[:80])  # batch grande pero acotado
            url = f"https://financialmodelingprep.com/stable/batch-quote-short?symbols={sym_csv}&apikey={FMP_KEY}"
            r = requests.get(url, timeout=15, headers=headers)
            if r.status_code == 200:
                payload = r.json()
                if isinstance(payload, list):
                    for item in payload:
                        symbol = str(item.get("symbol", "")).strip()
                        if not symbol:
                            continue
                        price = _to_float(item.get("price"), default=None)
                        if price is None:
                            continue
                        chg_pct = _to_float(item.get("changesPercentage"), default=None)
                        if chg_pct is None:
                            chg_pct = _to_float(item.get("change"), default=0.0)
                        result[symbol] = {"price": round(price, 2), "chg_pct": round(chg_pct, 2)}
        except Exception:
            pass

    # 2) Fallback individual para símbolos que no entraron por batch (índices, crypto, etc.)
    for sym in unique:
        if sym in result:
            continue
        try:
            enc = quote(sym, safe="")
            url = f"https://financialmodelingprep.com/stable/quote?symbol={enc}&apikey={FMP_KEY}"
            r = requests.get(url, timeout=12, headers=headers)
            if r.status_code != 200:
                continue
            payload = r.json()
            if not isinstance(payload, list) or not payload:
                continue
            item = payload[0]
            symbol = str(item.get("symbol", sym)).strip() or sym
            price = _to_float(item.get("price"), default=None)
            if price is None:
                continue
            chg_pct = _to_float(item.get("changesPercentage"), default=None)
            if chg_pct is None:
                chg_pct = _to_float(item.get("change"), default=0.0)
            result[symbol] = {"price": round(price, 2), "chg_pct": round(chg_pct, 2)}
        except Exception:
            continue
    return result


def _pick_quote(data: dict, candidates: list) -> tuple:
    for sym in candidates:
        if sym in data and data[sym].get("price") is not None:
            return data[sym], sym
    return {}, ""


def yf_quote_first(symbols: list) -> tuple:
    """Fallback Yahoo para cotizaciones actuales/cierre."""
    for sym in symbols:
        cached = get_cache(f"yfq_{sym}", ttl=900)
        if cached:
            return cached, sym
        try:
            h = yf.Ticker(sym).history(period="7d", interval="1d", auto_adjust=True)
            if h is None or h.empty or "Close" not in h:
                continue
            closes = h["Close"].dropna()
            if len(closes) == 0:
                continue
            last = float(closes.iloc[-1])
            prev = float(closes.iloc[-2]) if len(closes) >= 2 else last
            chg = round((last - prev) / prev * 100, 2) if prev else 0.0
            q = {"price": round(last, 2), "chg_pct": chg}
            set_cache(f"yfq_{sym}", q)
            return q, sym
        except Exception:
            continue
    return {}, ""


def yf_quote_bulk(symbols: list) -> dict:
    """Cotizaciones Yahoo en una sola descarga para reducir fallos/rate-limit."""
    out = {}
    syms = []
    for s in symbols:
        ss = (s or "").strip()
        if ss and ss not in syms:
            syms.append(ss)
    if not syms:
        return out
    try:
        data = yf.download(
            tickers=" ".join(syms),
            period="7d",
            interval="1d",
            auto_adjust=True,
            progress=False,
            threads=False,
            group_by="ticker",
        )
        if data is None or data.empty:
            return out

        # Caso 1 símbolo: columnas simples (Open/High/Low/Close...)
        if "Close" in data.columns:
            closes = data["Close"].dropna()
            if len(closes) >= 1:
                last = float(closes.iloc[-1])
                prev = float(closes.iloc[-2]) if len(closes) >= 2 else last
                chg = round((last - prev) / prev * 100, 2) if prev else 0.0
                out[syms[0]] = {"price": round(last, 2), "chg_pct": chg}
            return out

        # Caso múltiple: MultiIndex [ticker][field]
        for sym in syms:
            try:
                if sym not in data.columns.get_level_values(0):
                    continue
                sdf = data[sym]
                if "Close" not in sdf.columns:
                    continue
                closes = sdf["Close"].dropna()
                if len(closes) < 1:
                    continue
                last = float(closes.iloc[-1])
                prev = float(closes.iloc[-2]) if len(closes) >= 2 else last
                chg = round((last - prev) / prev * 100, 2) if prev else 0.0
                out[sym] = {"price": round(last, 2), "chg_pct": chg}
            except Exception:
                continue
    except Exception:
        return out
    return out


def av_quote_first(symbols: list) -> tuple:
    """Fallback Alpha Vantage para cotización."""
    for sym in symbols:
        q = av_quote_single(sym)
        if q and q.get("price") is not None:
            return q, sym
    return {}, ""


def yf_daily(symbol: str, timeseries: int = 130) -> list:
    """Fallback Yahoo para historial diario."""
    cached = get_cache(f"yfd_{symbol}", ttl=3600)
    if cached:
        return cached
    try:
        h = yf.Ticker(symbol).history(period="10mo", interval="1d", auto_adjust=True)
        if h is None or h.empty or "Close" not in h:
            return []
        out = []
        for idx, row in h.iterrows():
            c = _to_float(row.get("Close"), default=None)
            if c is None:
                continue
            out.append({"date": idx.strftime("%Y-%m-%d"), "close": round(c, 2)})
        if len(out) >= 5:
            return set_cache(f"yfd_{symbol}", out[-timeseries:])
        return []
    except Exception:
        return []


def fmp_daily(symbol: str, timeseries: int = 130) -> list:
    """Histórico diario FMP con fallback a Alpha Vantage."""
    cached = get_cache(f"fmpd_{symbol}", ttl=3600)
    if cached:
        return cached
    d912_hist = data912_daily(symbol, timeseries=timeseries)
    if len(d912_hist) >= 5:
        return d912_hist
    try:
        enc = quote(symbol, safe="")
        url = (
            f"https://financialmodelingprep.com/stable/historical-price-eod/full"
            f"?symbol={enc}&apikey={FMP_KEY}"
        )
        r = requests.get(url, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200:
            payload = r.json()
            if isinstance(payload, list):
                rows = payload
            elif isinstance(payload, dict):
                rows = payload.get("historical", [])
            else:
                rows = []
            hist = []
            for row in reversed(rows):
                date = row.get("date")
                close = _to_float(row.get("close"), default=None)
                if date and close is not None:
                    hist.append({"date": date, "close": round(close, 2)})
            if len(hist) >= 5:
                return set_cache(f"fmpd_{symbol}", hist[-timeseries:])
    except Exception:
        pass
    av_hist = av_daily(symbol)
    if len(av_hist) >= 5:
        return av_hist
    return yf_daily(symbol, timeseries=timeseries)


def dolarapi() -> dict:
    """Cotizaciones del dólar en Argentina."""
    cached = get_cache("dolar", ttl=300)
    if cached:
        return cached
    try:
        r    = requests.get("https://dolarapi.com/v1/dolares", timeout=8)
        data = r.json()
        tipos = {"oficial":"Oficial","blue":"Blue","contadoconliqui":"CCL","bolsa":"MEP"}
        result = {}
        for item in data:
            if item["casa"] in tipos:
                result[tipos[item["casa"]]] = {
                    "compra": item.get("compra", 0),
                    "venta":  item.get("venta", 0),
                }
        return set_cache("dolar", result)
    except Exception:
        return {}


# ═══════════════════════════════════════════════════════════════
#  SCANNER THREAD
# ═══════════════════════════════════════════════════════════════

def run_scanner_thread(session: str, params: dict = None, active_setups: list = None):
    global scanner_state
    scanner_state.update({"running":True,"session":session,"progress":0,
                          "results":[],"log":[],"current_ticker":""})
    try:
        import importlib, scanner as sc
        importlib.reload(sc)
        from config import CEDEARS, TOP_N, MIN_PROBABILITY
        mp = float(params.get("mp", MIN_PROBABILITY)) if params else MIN_PROBABILITY
        tn = int(params.get("tn", TOP_N))             if params else TOP_N
        scanner_state["total"] = len(CEDEARS)
        resultados = []
        for i, ticker in enumerate(CEDEARS, 1):
            scanner_state["progress"]       = i
            scanner_state["current_ticker"] = ticker
            try:
                res = sc.analizar_ticker(ticker, active_setups=active_setups)
                if res and res["probabilidad"] >= mp:
                    resultados.append(res)
                    scanner_state["log"].append(f"✅ {ticker.replace('.BA','')} — {res['setup']} | {res['probabilidad']}%")
                else:
                    scanner_state["log"].append(f"⬜ {ticker.replace('.BA','')} — sin setup")
            except Exception:
                scanner_state["log"].append(f"⚠️ {ticker} — error")
            time.sleep(0.3) if i % 10 != 0 else time.sleep(2)
        resultados.sort(key=lambda x:(x["probabilidad"],x["confluencias"],x["vol_rel"]),reverse=True)
        top = resultados[:tn]
        scanner_state["results"]  = top
        scanner_state["last_run"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        if top:
            sc.guardar_excel(top, session)
        sc.send_telegram(sc.build_telegram_message(top, session))
        scanner_state["log"].append(f"\n🏁 Completado. {len(top)} oportunidades.")
    except Exception as e:
        scanner_state["log"].append(f"❌ Error: {e}")
    finally:
        scanner_state["running"] = False


# ═══════════════════════════════════════════════════════════════
#  HISTORIAL
# ═══════════════════════════════════════════════════════════════

def get_historial():
    try:
        from openpyxl import load_workbook
        archivo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "registros_scanner.xlsx")
        if not os.path.exists(archivo): return []
        wb = load_workbook(archivo); ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                rows.append({
                    "fecha":row[0],"hora":row[1],"sesion":row[2],"ticker":row[3],
                    "setup":row[4],"confluencias":row[5],"probabilidad":row[6],
                    "precio":row[7],"target":row[8],"stop":row[9],
                    "rsi":row[10],"vol_rel":row[11],"atr":row[12],
                    "descripcion":row[13],
                    "resultado":row[14] if len(row)>14 else "Pendiente",
                })
        return list(reversed(rows))
    except Exception:
        return []


# ═══════════════════════════════════════════════════════════════
#  RUTAS
# ═══════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/run/<session>", methods=["POST"])
def run(session):
    session = session.upper()
    if session not in ["PRE-MARKET","APERTURA","CIERRE","MANUAL"]:
        return jsonify({"error":"Sesión inválida"}),400
    if scanner_state["running"]:
        return jsonify({"error":"El scanner ya está corriendo"}),400
    body = request.get_json(silent=True) or {}
    t = threading.Thread(target=run_scanner_thread,
        args=(session, body.get("params",{}), body.get("active_setups",None)), daemon=True)
    t.start()
    return jsonify({"ok":True,"session":session})


@app.route("/status")
def status():
    return jsonify({k: scanner_state[k] for k in
        ["running","session","progress","total","current_ticker","results","last_run",
         "log"]})


@app.route("/historial")
def historial():
    return jsonify(get_historial())


def build_market_payload() -> dict:
    cached = get_cache("market", ttl=240)
    if cached:
        return cached

    idx_defs = [
        {"n": "S&P 500", "s": "^GSPC", "fmp": ["^GSPC", "SPY"], "yf": ["^GSPC", "SPY"], "av": ["SPY"]},
        {"n": "NASDAQ", "s": "^IXIC", "fmp": ["^IXIC", "QQQ"], "yf": ["^IXIC", "QQQ"], "av": ["QQQ"]},
        {"n": "Dow Jones", "s": "^DJI", "fmp": ["^DJI", "DIA"], "yf": ["^DJI", "DIA"], "av": ["DIA"]},
        {"n": "Russell 2000", "s": "^RUT", "fmp": ["^RUT", "IWM"], "yf": ["^RUT", "IWM"], "av": ["IWM"]},
        {"n": "VIX", "s": "^VIX", "fmp": ["^VIX", "VXX"], "yf": ["^VIX", "VXX"], "av": ["VXX"]},
        {"n": "Bitcoin (BTC)", "s": "BTC", "fmp": ["BTCUSD", "IBIT", "MSTR"], "yf": ["BTC-USD", "IBIT", "MSTR"], "av": ["IBIT", "MSTR"]},
    ]
    com_defs = [
        {"n": "Oro", "fmp": ["GCUSD", "GLD"], "yf": ["GC=F", "GLD"], "av": ["GLD"]},
        {"n": "Petróleo WTI", "fmp": ["CLUSD", "USO"], "yf": ["CL=F", "USO"], "av": ["USO"]},
        {"n": "Plata", "fmp": ["SIUSD", "SLV"], "yf": ["SI=F", "SLV"], "av": ["SLV"]},
        {"n": "Cobre", "fmp": ["HGUSD", "CPER"], "yf": ["HG=F", "CPER"], "av": ["CPER"]},
    ]
    cry_defs = [
        {"n": "Bitcoin (BTC)", "fmp": ["BTCUSD", "IBIT", "MSTR"], "yf": ["BTC-USD", "IBIT", "MSTR"], "av": ["IBIT", "MSTR"]},
        {"n": "Ethereum (ETH)", "fmp": ["ETHUSD", "ETHA", "ETHE"], "yf": ["ETH-USD", "ETHA", "ETHE"], "av": ["ETHA", "ETHE"]},
    ]
    wl_syms = [
        "AAPL", "MSFT", "GOOGL", "AMZN", "NVDA", "META", "TSLA",
        "JPM", "BAC", "V", "MA", "XOM", "CVX", "JNJ", "UNH",
        "GGAL", "YPF", "MELI", "NU", "BABA",
    ]

    fmp_pool = []
    yf_pool = []
    for row in idx_defs + com_defs + cry_defs:
        fmp_pool.extend(row["fmp"])
        yf_pool.extend(row["yf"])
    yf_pool.extend(wl_syms)

    market_d912 = data912_quote_map(yf_pool)
    market_fmp = {}
    market_yf = {}
    movers_fmp = {}
    d912_hist_quote_cache = {}

    def d912_hist_quote_first(symbols: list) -> tuple:
        for sym in symbols:
            s = (sym or "").strip().upper()
            if not s:
                continue
            if s in d912_hist_quote_cache:
                q = d912_hist_quote_cache[s]
            else:
                h = data912_daily(s, timeseries=5)
                q = {}
                if len(h) >= 1:
                    last = _to_float(h[-1].get("close"), default=None)
                    prev = _to_float(h[-2].get("close"), default=last) if len(h) >= 2 else last
                    if last is not None:
                        chg = ((last - prev) / prev * 100) if prev else 0.0
                        q = {"price": round(last, 2), "chg_pct": round(chg, 2)}
                d912_hist_quote_cache[s] = q
            if q and q.get("price") is not None:
                return q, s
        return {}, ""

    def need_more_core_data() -> bool:
        for row in idx_defs + com_defs + cry_defs:
            q, _ = _pick_quote(market_d912, row["av"])
            if not q:
                return True
        return False

    need_core = need_more_core_data()
    need_movers = any((s not in market_d912 or market_d912[s].get("price") is None) for s in wl_syms)

    if need_core or need_movers:
        market_fmp = fmp_quote_safe(fmp_pool if need_core else wl_syms)
        movers_fmp = market_fmp if not need_core else fmp_quote_safe(wl_syms)

    if need_core:
        missing_after_d912_fmp = []
        for row in idx_defs + com_defs + cry_defs:
            q, _ = _pick_quote(market_d912, row["av"])
            if q:
                continue
            q, _ = _pick_quote(market_fmp, row["fmp"])
            if not q:
                missing_after_d912_fmp.extend(row["yf"])
        if missing_after_d912_fmp:
            market_yf = yf_quote_bulk(missing_after_d912_fmp)

    def pick_market(row):
        q, src = _pick_quote(market_d912, row["av"])
        if q:
            return q, f"D912:{src}"
        q, src = d912_hist_quote_first(row["av"])
        if q:
            return q, f"D912H:{src}"
        q, src = _pick_quote(market_fmp, row["fmp"])
        if q:
            return q, f"FMP:{src}"
        q, src = _pick_quote(market_yf, row["yf"])
        if q:
            return q, f"YF:{src}"
        q, src = av_quote_first(row["av"])
        if q:
            return q, f"AV:{src}"
        return {}, ""

    indices = []
    for row in idx_defs:
        q, src = pick_market(row)
        indices.append({"n": row["n"], "s": row["s"], "src": src, "d": q})

    commodities = []
    for row in com_defs:
        q, src = pick_market(row)
        commodities.append({"n": row["n"], "src": src, "d": q})

    crypto = []
    for row in cry_defs:
        q, src = pick_market(row)
        crypto.append({"n": row["n"], "src": src, "d": q})

    movers_data = {}
    for sym in wl_syms:
        if sym in market_d912 and market_d912[sym].get("price") is not None:
            movers_data[sym] = market_d912[sym]
            continue
        q, _ = d912_hist_quote_first([sym])
        if q:
            movers_data[sym] = q
            continue
        if sym in movers_fmp and movers_fmp[sym].get("price") is not None:
            movers_data[sym] = movers_fmp[sym]
            continue
        if sym in market_yf and market_yf[sym].get("price") is not None:
            movers_data[sym] = market_yf[sym]
            continue
        q, _ = av_quote_first([sym])
        if q:
            movers_data[sym] = q

    movers = []
    for s in wl_syms:
        q = movers_data.get(s, {})
        if q.get("chg_pct") is None:
            continue
        movers.append({"s": s, "price": q.get("price"), "chg_pct": q.get("chg_pct")})
    movers.sort(key=lambda x: x["chg_pct"], reverse=True)

    result = {
        "indices": indices,
        "commodities": [],
        "crypto": [],
        "gainers": movers[:5],
        "losers": movers[-5:][::-1],
        "dolar": dolarapi(),
    }

    last_ok = get_cache("market_last_ok", ttl=86400) or load_market_snapshot()

    def _merge_rows(cur_rows, old_rows, key_name):
        old_map = {}
        for r in (old_rows or []):
            k = r.get(key_name)
            if k:
                old_map[k] = r
        merged = []
        for r in (cur_rows or []):
            d = r.get("d", {}) if isinstance(r, dict) else {}
            has_price = isinstance(d, dict) and d.get("price") is not None
            if has_price:
                merged.append(r)
                continue
            old = old_map.get(r.get(key_name), {})
            old_d = old.get("d", {}) if isinstance(old, dict) else {}
            if isinstance(old_d, dict) and old_d.get("price") is not None:
                x = dict(r)
                x["d"] = old_d
                if not x.get("src"):
                    x["src"] = old.get("src", "SNAPSHOT")
                merged.append(x)
            else:
                merged.append(r)
        return merged

    if last_ok:
        result["indices"] = _merge_rows(result.get("indices", []), last_ok.get("indices", []), "n")
        if not result.get("gainers"):
            result["gainers"] = last_ok.get("gainers", [])
        if not result.get("losers"):
            result["losers"] = last_ok.get("losers", [])

    idx_ok = sum(1 for i in result["indices"] if i.get("d", {}).get("price") is not None)
    mover_ok = len(result.get("gainers", [])) + len(result.get("losers", []))
    # Umbral más flexible para no perder respaldo en días con proveedores incompletos.
    enough_for_snapshot = idx_ok >= 2 and mover_ok >= 6

    has_data = any(
        i.get("d", {}).get("price") is not None
        for i in result.get("indices", [])
    )
    has_data = has_data or len(result.get("gainers", [])) > 0 or len(result.get("losers", [])) > 0
    if has_data:
        set_cache("market", result)
        if enough_for_snapshot:
            set_cache("market_last_ok", result)
            save_market_snapshot(result)
        return result

    if last_ok:
        last_ok["dolar"] = result.get("dolar", {})
        return last_ok
    return result


@app.route("/market")
def market():
    """
    Endpoint unificado de datos de mercado.
    Retorna: indices, commodities, crypto, movers, dolar
    Usa FMP (indices/acciones) + dolarapi (dolar)
    """
    return jsonify(build_market_payload())

    cached = get_cache("market", ttl=900)
    if cached:
        return jsonify(cached)

    idx_defs = [
        {"n": "S&P 500", "s": "^GSPC", "candidates": ["^GSPC", "SPY"]},
        {"n": "NASDAQ", "s": "^IXIC", "candidates": ["^IXIC", "QQQ"]},
        {"n": "Dow Jones", "s": "^DJI", "candidates": ["^DJI", "DIA"]},
        {"n": "Russell 2000", "s": "^RUT", "candidates": ["^RUT", "IWM"]},
        {"n": "VIX", "s": "^VIX", "candidates": ["^VIX", "VXX"]},
        {"n": "Merval", "s": "^MERV", "candidates": ["^MERV", "ARGT"]},
    ]
    com_defs = [
        {"n": "Oro", "candidates": ["GCUSD", "GLD"]},
        {"n": "Petróleo WTI", "candidates": ["CLUSD", "USO"]},
        {"n": "Plata", "candidates": ["SIUSD", "SLV"]},
        {"n": "Cobre", "candidates": ["HGUSD", "CPER"]},
    ]
    cry_defs = [
        {"n": "Bitcoin (BTC)", "candidates": ["BTCUSD", "IBIT", "MSTR"]},
        {"n": "Ethereum (ETH)", "candidates": ["ETHUSD", "ETHA", "ETHE"]},
    ]

    pool = []
    for row in idx_defs + com_defs + cry_defs:
        pool.extend(row["candidates"])
    market_data = fmp_quote_safe(pool)
    idx_yf = {
        "S&P 500": ["^GSPC", "SPY"],
        "NASDAQ": ["^IXIC", "QQQ"],
        "Dow Jones": ["^DJI", "DIA"],
        "Russell 2000": ["^RUT", "IWM"],
        "VIX": ["^VIX", "VXX"],
        "Merval": ["^MERV", "ARGT"],
    }
    com_yf = {
        "Oro": ["GC=F", "GLD"],
        "PetrÃ³leo WTI": ["CL=F", "USO"],
        "Plata": ["SI=F", "SLV"],
        "Cobre": ["HG=F", "CPER"],
    }
    cry_yf = {
        "Bitcoin (BTC)": ["BTC-USD", "IBIT", "MSTR"],
        "Ethereum (ETH)": ["ETH-USD", "ETHE", "ETHA"],
    }
    idx_av = {
        "S&P 500": ["SPY"],
        "NASDAQ": ["QQQ"],
        "Dow Jones": ["DIA"],
        "Russell 2000": ["IWM"],
        "VIX": ["VXX"],
        "Merval": ["ARGT"],
    }
    com_av = {
        "Oro": ["GLD"],
        "Plata": ["SLV"],
        "Cobre": ["CPER"],
    }
    cry_av = {
        "Bitcoin (BTC)": ["IBIT", "MSTR"],
        "Ethereum (ETH)": ["ETHE", "ETHA"],
    }

    indices = []
    for row in idx_defs:
        q, src = _pick_quote(market_data, row["candidates"])
        if not q:
            q, ys = yf_quote_first(idx_yf.get(row["n"], []))
            if q:
                src = f"YF:{ys}"
        if not q:
            q, avs = av_quote_first(idx_av.get(row["n"], []))
            if q:
                src = f"AV:{avs}"
        indices.append({"n": row["n"], "s": row["s"], "src": src, "d": q})

    commodities = []
    for row in com_defs:
        q, src = _pick_quote(market_data, row["candidates"])
        if not q:
            q, ys = yf_quote_first(com_yf.get(row["n"], []))
            if q:
                src = f"YF:{ys}"
        if not q:
            av_list = com_av.get(row["n"], [])
            if not av_list and "Petr" in row["n"]:
                av_list = ["USO"]
            q, avs = av_quote_first(av_list)
            if q:
                src = f"AV:{avs}"
        commodities.append({"n": row["n"], "src": src, "d": q})

    crypto = []
    for row in cry_defs:
        q, src = _pick_quote(market_data, row["candidates"])
        if not q:
            q, ys = yf_quote_first(cry_yf.get(row["n"], []))
            if q:
                src = f"YF:{ys}"
        if not q:
            q, avs = av_quote_first(cry_av.get(row["n"], []))
            if q:
                src = f"AV:{avs}"
        crypto.append({"n": row["n"], "src": src, "d": q})

    wl_syms = ["AAPL","MSFT","GOOGL","AMZN","NVDA","META","TSLA",
               "JPM","BAC","V","MA","XOM","CVX","JNJ","UNH",
               "GGAL","YPF","MELI","NU","BABA"]
    wl_data = fmp_quote_safe(wl_syms)
    for s in wl_syms:
        if s not in wl_data or wl_data[s].get("price") is None:
            q, _ = yf_quote_first([s])
            if q:
                wl_data[s] = q
            else:
                q, _ = av_quote_first([s])
                if q:
                    wl_data[s] = q
    movers = [{"s": s, "price": wl_data[s]["price"], "chg_pct": wl_data[s]["chg_pct"]}
              for s in wl_syms if s in wl_data and wl_data[s].get("chg_pct") is not None]
    movers.sort(key=lambda x: x["chg_pct"], reverse=True)

    result = {
        "indices": indices,
        "commodities": commodities,
        "crypto": crypto,
        "gainers": movers[:5],
        "losers": movers[-5:][::-1],
        "dolar": dolarapi(),
    }
    set_cache("market", result)
    return jsonify(result)

    # ── Índices via FMP ──────────────────────────────────────
    idx_syms = ["^GSPC","^IXIC","^DJI","^RUT","^VIX","MERVAL"]
    idx_data = fmp_quote(idx_syms)

    indices = [
        {"n":"S&P 500",      "s":"^GSPC",  "d": idx_data.get("^GSPC",{})},
        {"n":"NASDAQ",       "s":"^IXIC",  "d": idx_data.get("^IXIC",{})},
        {"n":"Dow Jones",    "s":"^DJI",   "d": idx_data.get("^DJI",{})},
        {"n":"Russell 2000", "s":"^RUT",   "d": idx_data.get("^RUT",{})},
        {"n":"VIX",          "s":"^VIX",   "d": idx_data.get("^VIX",{})},
        {"n":"Merval",       "s":"MERVAL", "d": idx_data.get("MERVAL",{})},
    ]

    # ── Commodities via FMP ──────────────────────────────────
    com_syms  = ["GCUSD","CLUSD","SIUSD","HGUSD"]
    com_data  = fmp_quote(com_syms)
    commodities = [
        {"n":"Oro",          "d": com_data.get("GCUSD",{})},
        {"n":"Petróleo WTI", "d": com_data.get("CLUSD",{})},
        {"n":"Plata",        "d": com_data.get("SIUSD",{})},
        {"n":"Cobre",        "d": com_data.get("HGUSD",{})},
    ]

    # ── Crypto via FMP ───────────────────────────────────────
    cry_syms = ["BTCUSD","ETHUSD"]
    cry_data = fmp_quote(cry_syms)
    crypto = [
        {"n":"Bitcoin (BTC)",  "d": cry_data.get("BTCUSD",{})},
        {"n":"Ethereum (ETH)", "d": cry_data.get("ETHUSD",{})},
    ]

    # ── Movers via FMP ───────────────────────────────────────
    wl_syms = ["AAPL","MSFT","GOOGL","AMZN","NVDA","META","TSLA",
               "JPM","BAC","V","MA","XOM","CVX","JNJ","UNH",
               "GGAL","YPF","MELI","NU","BABA"]
    wl_data = fmp_quote(wl_syms)
    movers  = [{"s":s,"price":wl_data[s]["price"],"chg_pct":wl_data[s]["chg_pct"]}
               for s in wl_syms if s in wl_data and wl_data[s].get("chg_pct") is not None]
    movers.sort(key=lambda x: x["chg_pct"], reverse=True)

    # ── Dólar ────────────────────────────────────────────────
    dolar = dolarapi()

    result = {
        "indices":     indices,
        "commodities": commodities,
        "crypto":      crypto,
        "gainers":     movers[:5],
        "losers":      movers[-5:][::-1],
        "dolar":       dolar,
    }
    set_cache("market", result)
    return jsonify(result)


@app.route("/sector_data")
def sector_data():
    """Historial 6 meses via FMP (fallback a Alpha Vantage)."""
    tickers = [t.strip() for t in request.args.get("tickers","").split(",") if t.strip()]
    result  = {}
    for sym in tickers:
        cached = get_cache(f"sec_{sym}", ttl=3600)
        if cached:
            result[sym] = cached
            continue
        hist = fmp_daily(sym)
        if len(hist) < 5:
            continue
        last = hist[-1]["close"]
        prev = hist[-2]["close"] if len(hist)>=2 else last
        d = {"price":last,"chg_pct":round((last-prev)/prev*100,2),"history":hist}
        result[sym] = set_cache(f"sec_{sym}", d)
    return jsonify(result)


@app.route("/watchlist_data")
def watchlist_data():
    """RSI, MACD vía histórico diario FMP (fallback a Alpha Vantage)."""
    tickers = [t.strip() for t in request.args.get("tickers","").split(",") if t.strip()]
    if not tickers: return jsonify({})
    result = {}
    import pandas as pd
    for sym in tickers:
        try:
            hist = fmp_daily(sym)
            if len(hist) < 26: continue
            closes = pd.Series([h["close"] for h in hist])
            last    = float(closes.iloc[-1])
            prev    = float(closes.iloc[-2])
            chg_pct = round((last-prev)/prev*100,2)
            delta = closes.diff()
            gain  = delta.clip(lower=0).rolling(14).mean()
            loss  = (-delta.clip(upper=0)).rolling(14).mean()
            rs    = gain/loss.replace(0,float("nan"))
            rsi   = round(float(100-(100/(1+rs.iloc[-1]))),1)
            ema12 = closes.ewm(span=12,adjust=False).mean()
            ema26 = closes.ewm(span=26,adjust=False).mean()
            macd  = ema12-ema26
            sig   = macd.ewm(span=9,adjust=False).mean()
            mv,sv = round(float(macd.iloc[-1]),3),round(float(sig.iloc[-1]),3)
            result[sym] = {
                "price":last,"chg_pct":chg_pct,"rsi":rsi,
                "macd":mv,"macd_signal":sv,"macd_hist":round(mv-sv,3),
                "macd_str":"COMPRA" if mv>sv else "VENTA" if mv<sv else "NEUTRO",
                "atr":None,
                "sparkline":[round(float(v),2) for v in closes.iloc[-20:].tolist()],
            }
        except Exception:
            pass
        time.sleep(0.3)
    return jsonify(result)


@app.route("/news")
def news():
    """Noticias en español via RSS."""
    import xml.etree.ElementTree as ET
    cached = get_cache("news", ttl=300)
    if cached: return jsonify(cached)
    feeds = [
        ("El Cronista",  "https://www.cronista.com/files/rss/mercados.xml"),
        ("Ámbito",       "https://www.ambito.com/rss/pages/economia.html"),
        ("Infobae",      "https://www.infobae.com/feeds/rss/economia/"),
        ("iProfesional", "https://www.iprofesional.com/rss/home.xml"),
    ]
    items = []
    for source, url in feeds:
        try:
            fp = feedparser.parse(url)
            if fp.entries:
                for ent in fp.entries:
                    title = (ent.get("title") or "").strip()
                    link = (ent.get("link") or "").strip()
                    pub = (ent.get("published") or ent.get("updated") or "").strip()
                    if title and len(title) > 15:
                        items.append({"source":source,"title":title,"url":link,"time":pub[:16] if pub else ""})
                    if len(items)>=8:
                        break
                if len(items)>=8:
                    break
            r    = requests.get(url, timeout=6, headers={"User-Agent":"Mozilla/5.0"})
            root = ET.fromstring(r.content)
            for item in root.iter("item"):
                title = item.findtext("title","").strip()
                link  = item.findtext("link","").strip()
                pub   = item.findtext("pubDate","").strip()
                if title and len(title)>15:
                    items.append({"source":source,"title":title,"url":link,"time":pub[:16] if pub else ""})
                if len(items)>=8: break
        except Exception:
            pass
        if len(items)>=8: break
    if not items:
        items=[{"source":"Info","title":"No se pudieron cargar las noticias.","url":"#","time":""}]
    result = items[:5]
    set_cache("news", result)
    return jsonify(result)


@app.route("/health_data")
def health_data():
    """Estado rápido de proveedores de datos para diagnóstico."""
    out = {"ok": True}

    t0 = time.time()
    d912_rows = data912_live_usa()
    out["data912"] = {"rows": len(d912_rows), "ms": int((time.time() - t0) * 1000)}
    t0b = time.time()
    d912_spy_hist = data912_daily("SPY", timeseries=5)
    out["data912_hist_spy"] = {"rows": len(d912_spy_hist), "ms": int((time.time() - t0b) * 1000)}

    t1 = time.time()
    fmp = fmp_quote_safe(["SPY", "AAPL"])
    out["fmp"] = {"rows": len(fmp), "ms": int((time.time() - t1) * 1000)}

    t2 = time.time()
    yf_map = yf_quote_bulk(["SPY", "AAPL"])
    out["yahoo"] = {"rows": len(yf_map), "ms": int((time.time() - t2) * 1000)}

    t3 = time.time()
    av = av_quote_single("SPY")
    out["alpha_vantage"] = {"ok": bool(av), "ms": int((time.time() - t3) * 1000)}

    out["snapshot_exists"] = os.path.exists(MARKET_SNAPSHOT_PATH)
    out["market_cached"] = bool(get_cache("market", ttl=3600))
    return jsonify(out)


@app.route("/clear_cache")
def clear_cache():
    """Limpia el cache en memoria — llamar después de cambiar API keys."""
    _cache.clear()
    return jsonify({"ok": True, "message": "Cache limpiado"})


@app.route("/debug_fmp")
def debug_fmp():
    """Testea FMP directamente y muestra la respuesta cruda."""
    try:
        url = f"https://financialmodelingprep.com/stable/quote?symbol=AAPL&apikey={FMP_KEY}"
        r   = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        yq, ys = yf_quote_first(["AAPL"])
        return jsonify({
            "status":   r.status_code,
            "fmp_key":  FMP_KEY[:8] + "...",
            "response": r.json() if r.status_code == 200 else r.text[:500],
            "yf_fallback": {"symbol": ys, "quote": yq},
        })
    except Exception as e:
        return jsonify({"error": str(e)})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
